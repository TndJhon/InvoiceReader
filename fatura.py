import os
from openpyxl import Workbook
import pdfplumber
import re
from datetime import datetime

directory = 'pdf_invoices'
files = os.listdir(directory)  # Lista todos os arquivos disponíveis no diretório
files_quantity = len(files)     # Conta quantos arquivos há no diretório

if files_quantity == 0:  # Verifica a quantidade de arquivos; se for igual a zero, lança uma exceção
    raise Exception('Nenhum arquivo encontrado no diretório!')

wb = Workbook()           # Cria uma instância do Workbook
ws = wb.active            # Obtém a planilha ativa
ws.title = 'Aba Principal'  # Altera o nome da aba para 'Aba Principal'

ws['A1'] = 'Invoice #'    # Define o cabeçalho da coluna A como 'Invoice #'
ws['B1'] = 'Date'         # Define o cabeçalho da coluna B como 'Date'
ws['C1'] = 'File Name'    # Define o cabeçalho da coluna C como 'File Name'
ws['D1'] = 'Status'       # Define o cabeçalho da coluna D como 'Status'

last_empty_line = 1  # Inicializa a variável para encontrar a próxima linha vazia
while ws[f'A{last_empty_line}'].value is not None:  # Verifica se a célula na coluna A da linha atual não está vazia
    last_empty_line += 1  # Incrementa o contador de linhas até encontrar uma célula vazia

for file in files:  # Itera sobre cada arquivo no diretório
    with pdfplumber.open(directory + '/' + file) as pdf:  # Abre o arquivo PDF com pdfplumber
        first_page = pdf.pages[0]  # Obtém a primeira página do PDF
        pdf_text = first_page.extract_text()  # Extrai o texto da primeira página do PDF
        
    inv_number_re_pattern = r'INVOICE #(\d+)'  # Padrão regex para encontrar o número da fatura
    inv_date_re_pattern = r'DATE: (\d{2}/\d{2}/\d{4})'  # Padrão regex para encontrar a data da fatura
    
    match_number = re.search(inv_number_re_pattern, pdf_text)  # Busca o número da fatura no texto
    match_date = re.search(inv_date_re_pattern, pdf_text)  # Busca a data da fatura no texto
    
    if match_number:  # Se encontrar o número da fatura
        invoice_number = match_number.group(1)  # Captura o número da fatura
        ws[f'A{last_empty_line}'] = invoice_number  # Insere o número da fatura na célula apropriada
    else:
        ws[f'A{last_empty_line}'] = 'Número do invoice não localizado'  # Informa que o número não foi localizado    
        
    if match_date:  # Se encontrar a data da fatura
        invoice_date = match_date.group(1)  # Captura a data da fatura
        ws[f'B{last_empty_line}'] = invoice_date  # Insere a data da fatura na célula apropriada
    else:
        ws[f'B{last_empty_line}'] = "Data não localizada"  # Informa que a data não foi localizada
    
    ws[f'C{last_empty_line}'] = file  # Insere o nome do arquivo na célula apropriada
    ws[f'D{last_empty_line}'] = 'Sucesso!'  # Define o status como 'Sucesso'
    last_empty_line += 1  # Incrementa o contador de linhas para a próxima linha vazia

full_now = str(datetime.now()).replace(':', '-')  # Obtém a data e hora atuais e substitui os caracteres ':' por '-'
dot_index = full_now.index('.')  # Encontra o índice do caractere '.' na string de data e hora
now = full_now[:dot_index]  # Remove a parte fracionária dos segundos da string de data e hora

wb.save(f'Faturas - {now}.xlsx')  # Salva o arquivo Excel com o nome baseado na data e hora atuais


        


